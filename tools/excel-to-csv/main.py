import argparse
import csv
from openpyxl import load_workbook

def main():
    ap = argparse.ArgumentParser(description="Convert an Excel file to CSV")
    ap.add_argument("input", help="input .xlsx file")
    ap.add_argument("--sheet", default=None, help="sheet name (default: first)")
    ap.add_argument("-o", "--output", default=None, help="output .csv path")
    args = ap.parse_args()

    wb = load_workbook(args.input, read_only=True, data_only=True)
    ws = wb[args.sheet] if args.sheet else wb[wb.sheetnames[0]]
    out = args.output or f"{args.input.rsplit('.', 1)[0]}.csv"
    with open(out, "w", newline="", encoding="utf-8") as fh:
        writer = csv.writer(fh)
        for row in ws.iter_rows(values_only=True):
            writer.writerow(["" if v is None else v for v in row])
    wb.close()
    print(f"Wrote {out}")

if __name__ == "__main__":
    main()
