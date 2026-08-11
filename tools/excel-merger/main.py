import argparse
from pathlib import Path
from openpyxl import load_workbook, Workbook

def main():
    ap = argparse.ArgumentParser(description="Merge Excel files into one workbook")
    ap.add_argument("inputs", nargs="+", help="input .xlsx files")
    ap.add_argument("-o", "--output", default="merged.xlsx")
    ap.add_argument("--sheet-pattern", default=None, help="copy only sheets matching this substring")
    args = ap.parse_args()

    out = Workbook()
    out.remove(out.active)
    for path in args.inputs:
        wb = load_workbook(path, read_only=True, data_only=True)
        for sheet in wb.sheetnames:
            if args.sheet_pattern and args.sheet_pattern not in sheet:
                continue
            ws = out.create_sheet(f"{Path(path).stem}_{sheet}"[:31])
            for row in wb[sheet].iter_rows(values_only=True):
                ws.append(list(row))
        wb.close()
    out.save(args.output)
    print(f"Wrote {args.output} with {len(out.sheetnames)} sheets")

if __name__ == "__main__":
    main()
