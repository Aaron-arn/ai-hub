import argparse
import csv
from openpyxl import Workbook
from openpyxl.styles import Font

def main():
    ap = argparse.ArgumentParser(description="Generate an Excel workbook")
    ap.add_argument("output", help="output .xlsx path")
    ap.add_argument("--csv", help="input CSV file")
    ap.add_argument("--sheet", default="Sheet1")
    ap.add_argument("--header", action="store_true", help="bold first CSV row")
    ap.add_argument("--cells", help='inline cells, e.g. "A1=foo,B2=42"')
    args = ap.parse_args()

    wb = Workbook()
    ws = wb.active
    ws.title = args.sheet

    if args.cells:
        for pair in args.cells.split(","):
            coord, _, value = pair.partition("=")
            ws[coord] = value
    if args.csv:
        with open(args.csv, newline="", encoding="utf-8") as fh:
            for row in csv.reader(fh):
                ws.append(row)
        if args.header:
            for cell in ws[1]:
                cell.font = Font(bold=True)

    ws.auto_filter.ref = ws.dimensions
    wb.save(args.output)
    print(f"Wrote {args.output} ({ws.max_row} rows)")

if __name__ == "__main__":
    main()
